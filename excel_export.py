"""
Arsac Metal ERP — Excel Export Yardimcisi
Tum modullerden cagrilabilir.
"""
import os
from datetime import datetime


def excel_kaydet(parent, baslik, sutunlar, satirlar):
    """Excel veya CSV olarak kaydeder. openpyxl varsa xlsx, yoksa csv."""
    from PyQt5.QtWidgets import QFileDialog, QMessageBox
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        OPENPYXL = True
    except ImportError:
        OPENPYXL = False

    tarih_str = datetime.now().strftime("%Y%m%d_%H%M")
    temiz_baslik = "".join(c for c in baslik if c.isalnum() or c in " _-").strip()
    dosya_adi = "{}_{}.xlsx".format(temiz_baslik, tarih_str)

    yol, _ = QFileDialog.getSaveFileName(
        parent, "Excel Olarak Kaydet",
        os.path.join(os.path.expanduser("~"), "Desktop", dosya_adi),
        "Excel (*.xlsx)"
    )
    if not yol:
        return

    if not OPENPYXL:
        # openpyxl yoksa CSV olarak kaydet
        yol = yol.replace(".xlsx", ".csv")
        try:
            with open(yol, "w", encoding="utf-8-sig") as f:
                f.write(";".join(str(s) for s in sutunlar) + "\n")
                for satir in satirlar:
                    f.write(";".join(str(x) for x in satir) + "\n")
            QMessageBox.information(
                parent, "Kaydedildi",
                "CSV olarak kaydedildi (openpyxl yuklu degil):\n{}".format(yol)
            )
        except Exception as e:
            QMessageBox.critical(parent, "Hata", str(e))
        return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = temiz_baslik[:31]

        # Renkler
        HEADER_BG  = "2C3E50"
        HEADER_FG  = "FFFFFF"
        TITLE_BG   = "C0392B"
        TITLE_FG   = "FFFFFF"
        ROW_ALT    = "F8F9FA"
        BORDER_CLR = "DEE2E6"

        thin  = Side(style="thin",  color=BORDER_CLR)
        thick = Side(style="medium", color="AAAAAA")
        border_thin  = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

        # Baslik satiri
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=max(len(sutunlar), 1))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "{} — {} tarihli rapor".format(
            baslik, datetime.now().strftime("%d.%m.%Y %H:%M")
        )
        title_cell.font      = Font(name="Segoe UI", size=13, bold=True, color=TITLE_FG)
        title_cell.fill      = PatternFill("solid", fgColor=TITLE_BG)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Sutun basliklari
        for col, baslik_txt in enumerate(sutunlar, 1):
            c = ws.cell(row=2, column=col, value=str(baslik_txt))
            c.font      = Font(name="Segoe UI", size=11, bold=True, color=HEADER_FG)
            c.fill      = PatternFill("solid", fgColor=HEADER_BG)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = border_thin
        ws.row_dimensions[2].height = 22

        # Veri satirlari
        for r, satir in enumerate(satirlar, 3):
            bg = ROW_ALT if r % 2 == 0 else "FFFFFF"
            for col, deger in enumerate(satir, 1):
                c = ws.cell(row=r, column=col, value=deger)
                c.font      = Font(name="Segoe UI", size=10)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(vertical="center", wrap_text=False)
                c.border    = border_thin
            ws.row_dimensions[r].height = 18

        # Sutun genislikleri otomatik ayarla
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)

        # Dondur (baslik + sutun basligi)
        ws.freeze_panes = "A3"

        wb.save(yol)
        QMessageBox.information(
            parent, "Excel Kaydedildi",
            "Dosya basariyla kaydedildi:\n{}".format(yol)
        )

        # Dosyayi otomatik ac
        try:
            os.startfile(yol)
        except:
            pass

    except Exception as e:
        QMessageBox.critical(parent, "Excel Hatasi", str(e))