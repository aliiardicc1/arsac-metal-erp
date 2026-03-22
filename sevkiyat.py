"""
Arsac Metal ERP — Sevkiyat Modülü
Hazır siparişleri araç + şoför bilgisiyle sevke çıkarır.
"""
from styles import BTN_BLUE, BTN_GRAY, BTN_GREEN, BTN_ORANGE, BTN_PRIMARY, BTN_PURPLE, DIALOG_QSS, DURUM_RENK, GROUPBOX_QSS, INPUT, INPUT_QSS, LIST_QSS, SAYFA_QSS, TABLO_QSS, make_badge, make_buton, tab_qss, tablo_sag_tik_menu_ekle
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QColor
from datetime import datetime
try:
    from log import log_yaz
except:
    def log_yaz(c,n,i,d=""): pass
def excel_kaydet(*a, **kw):
    try:
        from excel_export import excel_kaydet as _ek
        _ek(*a, **kw)
    except ImportError:
        from PyQt5.QtWidgets import QMessageBox
        QMessageBox.warning(None, "Hata", "excel_export.py bulunamadi. Arsac_App klasorune kopyalayin.")
    except Exception as _e:
        from PyQt5.QtWidgets import QMessageBox
        QMessageBox.critical(None, "Excel Hatasi", str(_e))


class ParcaSevkDialog(QDialog):
    """Parça bazlı sevkiyat — araç ve şoför bilgisi giriş dialog'u."""
    def __init__(self, parca_adi, musteri, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Parça Sevk Et")
        self.setFixedSize(420, 300)
        self.setStyleSheet(DIALOG_QSS)
        self.plaka = ""; self.sofor = ""; self.telefon = ""
        self._build(parca_adi, musteri)

    def _build(self, parca_adi, musteri):
        lay = QVBoxLayout(self); lay.setContentsMargins(20,16,20,16); lay.setSpacing(10)

        t = QLabel("Parça Sevk Et"); t.setStyleSheet(
            "font-size:15px;font-weight:bold;color:#8e44ad;")
        lay.addWidget(t)

        info = QLabel("Parca: {}  |  Musteri: {}".format(parca_adi, musteri))
        info.setStyleSheet(
            "background:#f5eef8;border:1px solid #c39bd3;border-radius:6px;"
            "padding:8px;color:#6c3483;font-size:12px;")
        lay.addWidget(info)

        fg = QGridLayout(); fg.setSpacing(8)
        self.txt_plaka = QLineEdit(); self.txt_plaka.setPlaceholderText("Ornek: 16 ABC 123")
        self.txt_plaka.setFixedHeight(36)
        self.txt_sofor = QLineEdit(); self.txt_sofor.setPlaceholderText("Sofor adi soyadı")
        self.txt_sofor.setFixedHeight(36)
        self.txt_tel   = QLineEdit(); self.txt_tel.setPlaceholderText("Telefon (opsiyonel)")
        self.txt_tel.setFixedHeight(36)

        for row, (lbl, w) in enumerate([
            ("Plaka *:", self.txt_plaka),
            ("Sofor *:", self.txt_sofor),
            ("Telefon:", self.txt_tel),
        ]):
            fg.addWidget(QLabel(lbl), row, 0)
            fg.addWidget(w, row, 1)
        lay.addLayout(fg)

        bh = QHBoxLayout(); bh.addStretch()
        bi = QPushButton("Iptal")
        bi.setStyleSheet("background:#dcdde1;color:#2c3e50;border-radius:7px;padding:7px 16px;font-weight:bold;")
        bi.clicked.connect(self.reject)
        bk = QPushButton("Sevk Et")
        bk.setFixedHeight(38)
        bk.setStyleSheet("background:#8e44ad;color:white;border-radius:7px;padding:7px 20px;font-weight:bold;font-size:13px;border:none;")
        bk.clicked.connect(self._kaydet)
        bh.addWidget(bi); bh.addWidget(bk); lay.addLayout(bh)

    def _kaydet(self):
        plaka = self.txt_plaka.text().strip()
        sofor = self.txt_sofor.text().strip()
        if not plaka or not sofor:
            QMessageBox.warning(self, "Eksik", "Plaka ve sofor zorunlu!"); return
        self.plaka = plaka; self.sofor = sofor
        self.telefon = self.txt_tel.text().strip()
        self.accept()


class YeniSevkDialog(QDialog):
    def __init__(self, cursor, conn, parent=None):
        super().__init__(parent)
        self.cursor = cursor
        self.conn   = conn
        self.setWindowTitle("Yeni Sevkiyat Olustur")
        self.setMinimumSize(640, 560)
        self.setStyleSheet(DIALOG_QSS)
        self.init_ui()

    def init_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20,16,20,16); lay.setSpacing(12)

        # Araç bilgileri
        arac_box = QGroupBox("🚛 Araç & Şoför Bilgileri")
        ag = QGridLayout(arac_box); ag.setSpacing(8)

        def _le(ph):
            w = QLineEdit(); w.setPlaceholderText(ph); w.setFixedHeight(40); return w

        self.txt_plaka   = _le("34 ABC 123")
        self.txt_sofor   = _le("Şoför adı soyadı")
        self.txt_telefon = _le("Şoför telefonu")
        self.txt_notlar  = _le("Notlar (opsiyonel)")

        ag.addWidget(QLabel("Plaka:"),    0,0); ag.addWidget(self.txt_plaka,   0,1)
        ag.addWidget(QLabel("Şoför:"),    0,2); ag.addWidget(self.txt_sofor,   0,3)
        ag.addWidget(QLabel("Telefon:"),  1,0); ag.addWidget(self.txt_telefon, 1,1)
        ag.addWidget(QLabel("Notlar:"),   1,2); ag.addWidget(self.txt_notlar,  1,3)
        lay.addWidget(arac_box)

        # Hazır siparişler
        sip_box = QGroupBox("📦 Sevk Edilecek Siparişler (Hazır olanlar)")
        sv = QVBoxLayout(sip_box)

        self.tablo = QTableWidget(0,5)
        self.tablo.setHorizontalHeaderLabels(["✓","SİPARİŞ NO","MÜŞTERİ","TERMİN","TOPLAM"])
        self.tablo.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tablo.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.tablo.setColumnWidth(0, 40)
        self.tablo.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tablo.verticalHeader().setVisible(False)
        self.tablo.setShowGrid(False)
        self.tablo.setAlternatingRowColors(True)
        self.tablo.setStyleSheet("QTableWidget{background:white;border:none;font-size:13px;} QHeaderView::section{background:#f4f6f8;color:#2c3e50;padding:8px;font-weight:bold;border:none;border-bottom:2px solid #dfe6e9;}")
        sv.addWidget(self.tablo)

        lbl_bilgi = QLabel("💡 Sevk etmek istediğiniz siparişleri işaretleyin")
        lbl_bilgi.setStyleSheet("color:#7f8c8d;font-size:11px;font-weight:normal;margin-top:4px;")
        sv.addWidget(lbl_bilgi)
        lay.addWidget(sip_box)

        # Butonlar
        bl = QHBoxLayout()
        btn_iptal = QPushButton("İptal"); btn_iptal.setStyleSheet("background:#dcdde1;color:#2c3e50;border-radius:8px;padding:10px 24px;font-weight:bold;")
        btn_iptal.clicked.connect(self.reject)
        btn_kaydet = QPushButton("Sevkiyat Olustur")
        btn_kaydet.setStyleSheet("background:#e67e22;color:white;border-radius:8px;padding:10px 24px;font-weight:bold;font-size:14px;")
        btn_kaydet.clicked.connect(self._kaydet)
        bl.addWidget(btn_iptal); bl.addStretch(); bl.addWidget(btn_kaydet)
        lay.addLayout(bl)

        self._hazir_siparisleri_yukle()

    def _hazir_siparisleri_yukle(self):
        try:
            self.cursor.execute("""
                SELECT id, sip_no, musteri, termin, genel_toplam
                FROM siparisler WHERE durum='Hazır' ORDER BY id DESC
            """)
            self.tablo.setRowCount(0)
            for i, (sid, sno, mus, ter, top) in enumerate(self.cursor.fetchall()):
                self.tablo.insertRow(i)
                chk = QCheckBox(); chk.setStyleSheet("margin-left:10px;")
                self.tablo.setCellWidget(i, 0, chk)
                for j, v in enumerate([sno, mus or "-", ter or "-", f"{float(top or 0):,.2f} ₺"]):
                    item = QTableWidgetItem(v); item.setTextAlignment(Qt.AlignCenter)
                    item.setData(Qt.UserRole, sid)
                    self.tablo.setItem(i, j+1, item)
        except Exception as e:
            print(f"Hazır sipariş yükleme hatası: {e}")

    def _kaydet(self):
        plaka = self.txt_plaka.text().strip()
        sofor = self.txt_sofor.text().strip()
        if not plaka or not sofor:
            QMessageBox.warning(self,"Hata","Plaka ve şoför bilgisi zorunludur!"); return

        # Seçili siparişleri bul
        secili = []
        for r in range(self.tablo.rowCount()):
            chk = self.tablo.cellWidget(r, 0)
            if chk and chk.isChecked():
                item = self.tablo.item(r, 1)
                if item: secili.append(item.data(Qt.UserRole))

        if not secili:
            QMessageBox.warning(self,"Hata","En az bir sipariş seçmelisiniz!"); return

        try:
            tarih = datetime.now().strftime("%d.%m.%Y %H:%M")
            sip_listesi = ", ".join(str(s) for s in secili)

            self.cursor.execute("""
                INSERT INTO sevkiyatlar (plaka, sofor, telefon, tarih, siparis_listesi, notlar, durum)
                VALUES (?,?,?,?,?,?,'Yolda')
            """, (plaka, sofor, self.txt_telefon.text().strip(), tarih,
                  sip_listesi, self.txt_notlar.text().strip()))
            sev_id = self.cursor.lastrowid

            for sid in secili:
                self.cursor.execute("INSERT INTO sevkiyat_siparisler (sevkiyat_id, siparis_id) VALUES (?,?)", (sev_id, sid))
                self.cursor.execute("UPDATE siparisler SET durum='Sevk Edildi', arac=?, sofor=? WHERE id=?",
                                    (plaka, sofor, sid))

            self.conn.commit()
            log_yaz(self.cursor, self.conn, "SEVKIYAT_OLUSTURULDU",
                    f"{plaka} | {sofor} | {len(secili)} siparis")

            # Wolvox aktarım sorusu
            cevap = QMessageBox.question(
                self, "Sevkiyat Olusturuldu",
                "{} arac, {} sofor ile {} siparis sevk edildi.\n\n"
                "Wolvox irsaliyesi olusturmak ister misiniz?".format(
                    plaka, sofor, len(secili)),
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No)

            if cevap == QMessageBox.Yes:
                self._wolvox_irsaliye_sev(sev_id, secili)

            self.accept()
        except Exception as e:
            QMessageBox.critical(self,"Hata",str(e))

    def _wolvox_irsaliye_sev(self, sev_id, siparis_idler):
        """Sevkiyat sonrasi Wolvox irsaliye dialog'u acar."""
        # Parcalari veritabanindan al
        parcalar = []
        try:
            for sid in siparis_idler:
                # Once siparis bilgisini al
                self.cursor.execute(
                    "SELECT sip_no, musteri FROM siparisler WHERE id=?", (sid,))
                sip_row = self.cursor.fetchone()
                sip_no  = (sip_row[0] if isinstance(sip_row, (list,tuple)) else sip_row.get("sip_no","")) if sip_row else str(sid)
                musteri = (sip_row[1] if isinstance(sip_row, (list,tuple)) else sip_row.get("musteri","")) if sip_row else ""

                self.cursor.execute("""
                    SELECT urun_adi, adet, kg, malzeme
                    FROM siparis_kalemleri WHERE siparis_id=?
                """, (sid,))
                for row in self.cursor.fetchall():
                    urun    = (row[0] if isinstance(row,(list,tuple)) else row.get("urun_adi","")) or ""
                    adet    = (row[1] if isinstance(row,(list,tuple)) else row.get("adet",1)) or 1
                    kg      = (row[2] if isinstance(row,(list,tuple)) else row.get("kg",0)) or 0
                    malzeme = (row[3] if isinstance(row,(list,tuple)) else row.get("malzeme","")) or ""
                    parcalar.append({
                        "parca":   urun or malzeme or "-",
                        "adet":    float(adet),
                        "kg":      float(kg),
                        "sip_no":  sip_no,
                        "musteri": musteri,
                    })
        except Exception as e:
            print("Parca yukle hatasi:", e)

        if not parcalar:
            QMessageBox.information(self, "Bilgi", "Aktarilacak kalem bulunamadi.")
            return

        # Detay dialog'u ac
        dlg = WolvoxIrsaliyeDialog(parcalar, self)
        if dlg.exec_() == QDialog.Accepted:
            self._wolvox_excel_olustur(dlg.get_data())


    def _wolvox_excel_olustur(self, data):
        """Onaylanan veriyi Wolvox Excel formatinda kaydeder."""
        from PyQt5.QtWidgets import QFileDialog
        import os

        dosya, _ = QFileDialog.getSaveFileName(
            self, "Wolvox Irsaliye Kaydet",
            "Wolvox_Irsaliye_{}.xlsx".format(
                __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M")),
            "Excel Dosyasi (*.xlsx)")
        if not dosya:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Temp1"

            basliklar = [
                "Sira No", "Stok Kodu", "Stok Adi", "Miktari", "Birimi",
                "Temel Mik.", "Temel Brm.", "Fiyati", "Birim 2 Fiyati",
                "KDV Durumu", "KDV Hrc.Fiyat", "Ara Tutari", "KDV",
                "KDV siz Toplam", "KDV li Toplam", "Vade Gunu", "Ek Bilgi 1", "Depo Adi"
            ]

            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill("solid", start_color="2C3E50")
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin      = Side(style="thin", color="DDDDDD")
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, baslik in enumerate(basliklar, 1):
                cell = ws.cell(row=1, column=col, value=baslik)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = hdr_align; cell.border = border
            ws.row_dimensions[1].height = 24

            genislikler = [8,14,30,10,8,10,8,12,12,10,12,12,8,14,14,10,20,14]
            for i, g in enumerate(genislikler, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = g

            alt_font  = Font(size=11)
            alt_align = Alignment(horizontal="center", vertical="center")

            for idx, p in enumerate(data, 1):
                row  = idx + 1
                fill = PatternFill("solid", start_color="FFFFFF" if idx%2==1 else "F8F9FA")
                degerler = [
                    idx,
                    p["stok_kodu"],
                    p["parca"],
                    p["miktar"],
                    p["birim"],
                    "", "", 0, 0,
                    "Dahil", 0, 0, 20, 0, 0, 0,
                    "{} - {}".format(p["sip_no"], p["musteri"]),
                    "Merkez",
                ]
                for col, val in enumerate(degerler, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = alt_font; cell.alignment = alt_align
                    cell.fill = fill; cell.border = border
                ws.row_dimensions[row].height = 22

            wb.save(dosya)
            cevap = QMessageBox.question(
                self, "Basarili",
                "{} kalem aktarildi!\n\nDosya acilsin mi?".format(len(data)),
                QMessageBox.Yes | QMessageBox.No)
            if cevap == QMessageBox.Yes:
                try:
                    import sys
                    if sys.platform == "win32": os.startfile(dosya)
                except: pass

        except ImportError:
            QMessageBox.critical(self, "Hata",
                "openpyxl bulunamadi!\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Hata", "Excel olusturulamadi:\n{}".format(e))


class WolvoxIrsaliyeDialog(QDialog):
    """
    Wolvox irsaliyesi hazirlanmadan once her kalem icin:
    - Stok Kodu girisi
    - Adet mi / KG mi secimi
    sorar.
    """
    def __init__(self, parcalar, parent=None):
        super().__init__(parent)
        self.parcalar = parcalar
        self.satirlar = []  # (stok_kodu_widget, miktar_widget, birim_widget)
        self.setWindowTitle("Wolvox Irsaliye Hazirla")
        self.setMinimumWidth(820)
        self.setMinimumHeight(500)
        self.setStyleSheet("""
            QDialog{background:#f4f6f9;}
            QLabel{color:#2c3e50;font-size:13px;}
            QLineEdit,QComboBox{border:1.5px solid #dcdde1;border-radius:6px;
                padding:5px 8px;background:white;color:#2c3e50;font-size:13px;}
            QLineEdit:focus{border:1.5px solid #2980b9;}
            QHeaderView::section{background:#2c3e50;color:white;padding:7px;
                font-weight:bold;font-size:12px;border:none;}
            QTableWidget{background:white;border:1px solid #dcdde1;
                border-radius:8px;gridline-color:#f0f2f5;}
            QTableWidget::item{padding:4px 6px;color:#2c3e50;}
        """)
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 16, 20, 16)
        lay.setSpacing(12)

        # Baslik
        hdr = QHBoxLayout()
        lbl = QLabel("Wolvox Irsaliye Kalemleri")
        lbl.setStyleSheet("font-size:16px;font-weight:bold;color:#2c3e50;")
        hdr.addWidget(lbl)
        hdr.addStretch()

        # Toplu birim secimi
        lbl_toplu = QLabel("Hepsine Uygula:")
        lbl_toplu.setStyleSheet("font-size:12px;color:#7f8c8d;")
        hdr.addWidget(lbl_toplu)
        self.cmb_toplu = QComboBox()
        self.cmb_toplu.addItems(["Adet", "Kg"])
        self.cmb_toplu.setFixedWidth(80)
        self.cmb_toplu.setFixedHeight(32)
        self.cmb_toplu.currentTextChanged.connect(self._toplu_birim)
        hdr.addWidget(self.cmb_toplu)
        lay.addLayout(hdr)

        # Tablo
        self.tablo = QTableWidget(len(self.parcalar), 5)
        self.tablo.setHorizontalHeaderLabels([
            "Stok Adi", "Siparis / Musteri", "Stok Kodu", "Miktar", "Birim (Adet/Kg)"
        ])
        self.tablo.verticalHeader().setVisible(False)
        self.tablo.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tablo.setSelectionMode(QTableWidget.NoSelection)
        self.tablo.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tablo.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.tablo.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.tablo.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.tablo.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.tablo.verticalHeader().setDefaultSectionSize(40)
        self.tablo.setAlternatingRowColors(True)

        INP = "border:1.5px solid #dcdde1;border-radius:6px;padding:4px 8px;background:white;color:#2c3e50;font-size:13px;"

        for i, p in enumerate(self.parcalar):
            # Stok Adi (salt okunur)
            it0 = QTableWidgetItem(p["parca"])
            it0.setFlags(Qt.ItemIsEnabled)
            self.tablo.setItem(i, 0, it0)

            # Siparis / Musteri
            it1 = QTableWidgetItem("{} - {}".format(p["sip_no"], p["musteri"]))
            it1.setFlags(Qt.ItemIsEnabled)
            it1.setForeground(QColor("#7f8c8d"))
            self.tablo.setItem(i, 1, it1)

            # Stok Kodu girisi
            txt_kod = QLineEdit()
            txt_kod.setPlaceholderText("Stok kodu girin...")
            txt_kod.setFixedHeight(32)
            txt_kod.setStyleSheet(INP)
            self.tablo.setCellWidget(i, 2, txt_kod)

            # Miktar
            txt_mik = QLineEdit()
            txt_mik.setFixedHeight(32)
            txt_mik.setStyleSheet(INP)
            txt_mik.setFixedWidth(80)
            txt_mik.setPlaceholderText("0")
            # Varsayilan: adet
            txt_mik.setText("{:g}".format(p["adet"]))
            self.tablo.setCellWidget(i, 3, txt_mik)

            # Birim secimi
            cmb = QComboBox()
            cmb.addItems(["Adet", "Kg"])
            cmb.setFixedHeight(32)
            cmb.setFixedWidth(80)
            # Kg dolu ise Kg'a sec
            if float(p.get("kg", 0) or 0) > 0:
                cmb.setCurrentText("Kg")
                txt_mik.setText("{:g}".format(float(p["kg"])))
            # Birim degisince miktari guncelle
            def _birim_degisti(birim, mik_w=txt_mik, parca=p):
                if birim == "Kg":
                    mik_w.setText("{:g}".format(float(parca.get("kg",0) or 0)))
                else:
                    mik_w.setText("{:g}".format(float(parca.get("adet",1) or 1)))
            cmb.currentTextChanged.connect(_birim_degisti)
            self.tablo.setCellWidget(i, 4, cmb)

            self.satirlar.append((txt_kod, txt_mik, cmb))

        lay.addWidget(self.tablo)

        # Bilgi notu
        not_lbl = QLabel("* Stok kodu bos birakilabilir, Wolvox'ta eslestirebilirsiniz.")
        not_lbl.setStyleSheet("font-size:11px;color:#95a5a6;")
        lay.addWidget(not_lbl)

        # Butonlar
        btn_lay = QHBoxLayout()
        btn_lay.addStretch()
        btn_iptal = QPushButton("Iptal")
        btn_iptal.setFixedHeight(38)
        btn_iptal.setStyleSheet("background:#dcdde1;color:#2c3e50;border-radius:8px;"
                                "padding:6px 20px;font-weight:bold;border:none;")
        btn_iptal.clicked.connect(self.reject)

        btn_olustur = QPushButton("Excel Olustur")
        btn_olustur.setFixedHeight(38)
        btn_olustur.setStyleSheet("background:#8e44ad;color:white;border-radius:8px;"
                                  "padding:6px 24px;font-weight:bold;font-size:13px;border:none;")
        btn_olustur.clicked.connect(self.accept)
        btn_lay.addWidget(btn_iptal)
        btn_lay.addWidget(btn_olustur)
        lay.addLayout(btn_lay)

    def _toplu_birim(self, birim):
        """Tum satirlara ayni birimi uygula."""
        for i, (_, mik_w, cmb) in enumerate(self.satirlar):
            cmb.blockSignals(True)
            cmb.setCurrentText(birim)
            cmb.blockSignals(False)
            p = self.parcalar[i]
            if birim == "Kg":
                mik_w.setText("{:g}".format(float(p.get("kg",0) or 0)))
            else:
                mik_w.setText("{:g}".format(float(p.get("adet",1) or 1)))

    def get_data(self):
        """Dialog onaylaninca hazirlanmis veriyi dondur."""
        sonuc = []
        for i, (kod_w, mik_w, cmb) in enumerate(self.satirlar):
            p = self.parcalar[i]
            try:
                mik = float(mik_w.text().replace(",",".")) if mik_w.text().strip() else 0
            except:
                mik = 0
            sonuc.append({
                "parca":     p["parca"],
                "stok_kodu": kod_w.text().strip(),
                "miktar":    mik,
                "birim":     cmb.currentText(),
                "sip_no":    p["sip_no"],
                "musteri":   p["musteri"],
            })
        return sonuc


class SevkiyatSayfasi(QWidget):
    def __init__(self, cursor, conn, user_role):
        super().__init__()
        self.cursor    = cursor
        self.conn      = conn
        self.user_role = user_role
        self.init_ui()
        self.yenile()

    def init_ui(self):
        self.setStyleSheet(SAYFA_QSS + INPUT_QSS + TABLO_QSS)
        lay = QVBoxLayout(self); lay.setContentsMargins(24,16,24,16); lay.setSpacing(14)

        # Üst
        ust = QHBoxLayout()
        lbl = QLabel("SEVKIYAT")
        lbl.setStyleSheet("font-size:18px;font-weight:bold;color:#2c3e50;")
        ust.addWidget(lbl); ust.addStretch()

        self.k_parca  = self._kart("BEKLEYEN PARCA", "0", "#8e44ad")
        self.k_hazir  = self._kart("SEVKE HAZIR",    "0", "#e67e22")
        self.k_yolda  = self._kart("YOLDA",          "0", "#f39c12")
        self.k_teslim = self._kart("TESLIM",         "0", "#27ae60")
        for k in [self.k_parca, self.k_hazir, self.k_yolda, self.k_teslim]:
            ust.addWidget(k)
        ust.addSpacing(10)

        if self.user_role != "readonly":
            btn_yeni = QPushButton("Yeni Sevkiyat")
            btn_yeni.setFixedHeight(38)
            btn_yeni.setStyleSheet("background:#e67e22;color:white;border-radius:8px;font-weight:bold;font-size:13px;padding:4px 16px;border:none;")
            btn_yeni.clicked.connect(self._yeni_sevk)
            ust.addWidget(btn_yeni)

        btn_excel_sev = QPushButton("Excel")
        btn_excel_sev.setFixedHeight(38)
        btn_excel_sev.setStyleSheet("background:#27ae60;color:white;border-radius:8px;font-weight:bold;font-size:12px;padding:4px 14px;border:none;")
        btn_excel_sev.clicked.connect(self._excel_export)
        ust.addWidget(btn_excel_sev)

        btn_wolvox = QPushButton("📄 Wolvox İrsaliyesi")
        btn_wolvox.setFixedHeight(38)
        btn_wolvox.setStyleSheet("background:#8e44ad;color:white;border-radius:8px;font-weight:bold;font-size:12px;padding:4px 14px;border:none;")
        btn_wolvox.clicked.connect(self._wolvox_irsaliye)
        ust.addWidget(btn_wolvox)

        btn_yenile = QPushButton("Yenile")
        btn_yenile.setFixedHeight(38)
        btn_yenile.setStyleSheet("background:#dcdde1;color:#2c3e50;border-radius:8px;font-weight:bold;font-size:12px;padding:4px 14px;border:none;")
        btn_yenile.clicked.connect(self.yenile)
        ust.addWidget(btn_yenile)
        lay.addLayout(ust)

        # Sekmeler
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane{border:1px solid #dcdde1;border-radius:10px;background:white;
            color: #2c3e50;}
            QTabBar::tab{background:#ecf0f1;color:#2c3e50;padding:8px 20px;
                         border-radius:6px 6px 0 0;font-weight:bold;}
            QTabBar::tab:selected{background:#e67e22;color:white;}
        """)

        # ── Sekme 1: Sipariş Bazlı Üretimden Gelenler ──────────────
        t1 = QWidget(); t1l = QVBoxLayout(t1)
        t1l.setContentsMargins(10,10,10,10); t1l.setSpacing(6)

        # Üst araç çubuğu
        t1h = QHBoxLayout(); t1h.setSpacing(6)

        # Arama kutusu
        self.txt_ara_parca = QLineEdit()
        self.txt_ara_parca.setPlaceholderText("🔍  Siparis no, musteri veya parca adi ara...")
        self.txt_ara_parca.setFixedHeight(32)
        self.txt_ara_parca.setStyleSheet(
            "border:1.5px solid #dcdde1;border-radius:7px;padding:5px 10px;"
            "background:white;color:#2c3e50;font-size:12px;")
        self.txt_ara_parca.textChanged.connect(self._filtrele_parca)
        t1h.addWidget(self.txt_ara_parca)

        # Filtre: tüm siparişler veya tek sipariş
        self.cmb_sip_filtre = QComboBox()
        self.cmb_sip_filtre.addItem("Tum Siparisler")
        self.cmb_sip_filtre.setFixedHeight(32)
        self.cmb_sip_filtre.setFixedWidth(200)
        self.cmb_sip_filtre.setStyleSheet(
            "border:1.5px solid #dcdde1;border-radius:6px;padding:4px 8px;"
            "background:white;color:#2c3e50;font-size:12px;")
        self.cmb_sip_filtre.currentTextChanged.connect(self._sip_filtre_degisti)
        t1h.addWidget(self.cmb_sip_filtre)

        btn_tumunu_sec = QPushButton("Tumunu Sec")
        btn_tumunu_sec.setFixedHeight(32)
        btn_tumunu_sec.setStyleSheet(
            "background:#ecf0f1;color:#2c3e50;border-radius:6px;"
            "padding:5px 12px;font-size:12px;font-weight:bold;border:1px solid #dcdde1;")
        btn_tumunu_sec.clicked.connect(self._tumunu_sec)

        self.btn_secili_sevk = QPushButton("Secilileri Sevk Et")
        self.btn_secili_sevk.setFixedHeight(32)
        self.btn_secili_sevk.setStyleSheet(
            "background:#8e44ad;color:white;border-radius:6px;"
            "padding:5px 14px;font-size:12px;font-weight:bold;border:none;")
        self.btn_secili_sevk.clicked.connect(self._secilileri_sevk_et)
        self.btn_secili_sevk.setEnabled(False)

        t1h.addWidget(btn_tumunu_sec)
        t1h.addWidget(self.btn_secili_sevk)
        t1l.addLayout(t1h)

        # Ana tablo: sipariş gruplu
        self.tablo_parca = QTableWidget(0, 7)
        tablo_sag_tik_menu_ekle(self.tablo_parca)
        self.tablo_parca.setHorizontalHeaderLabels(
            ["", "Siparis No", "Musteri", "Parca Adi", "Adet", "Tamamlanan", "Tarih"])
        self.tablo_parca.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.tablo_parca.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.tablo_parca.setColumnWidth(0, 36)
        for c in [1,2,4,5,6]:
            self.tablo_parca.horizontalHeader().setSectionResizeMode(c, QHeaderView.ResizeToContents)
        self.tablo_parca.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tablo_parca.verticalHeader().setVisible(False)
        self.tablo_parca.setShowGrid(True)
        self.tablo_parca.setAlternatingRowColors(True)
        self.tablo_parca.verticalHeader().setDefaultSectionSize(40)
        self.tablo_parca.setSelectionBehavior(QTableWidget.SelectRows)
        self.tablo_parca.setStyleSheet("""
            QTableWidget{background:white;color:#2c3e50;gridline-color:#f0f2f5;}
            QTableWidget::item{color:#2c3e50;padding:5px;}
            QTableWidget::item:selected{background:#8e44ad;color:white;}
            QHeaderView::section{background:#2c3e50;color:white;padding:8px;
                font-weight:bold;border:none;border-right:1px solid #3d5166;}
            QTableWidget::item:alternate{background:#f8f9fa;}
        """)
        t1l.addWidget(self.tablo_parca)
        self.tabs.addTab(t1, "Uretimden Gelenler")

        # Sekme 2: Hazır Siparişler
        t2 = QWidget(); t2l = QVBoxLayout(t2); t2l.setContentsMargins(10,10,10,10); t2l.setSpacing(6)
        t2h = QHBoxLayout()
        self.txt_ara_hazir = QLineEdit()
        self.txt_ara_hazir.setPlaceholderText("🔍  Siparis no veya musteri ara...")
        self.txt_ara_hazir.setFixedHeight(32)
        self.txt_ara_hazir.setStyleSheet(
            "border:1.5px solid #dcdde1;border-radius:7px;padding:5px 10px;"
            "background:white;color:#2c3e50;font-size:12px;")
        self.txt_ara_hazir.textChanged.connect(self._filtrele_hazir)
        t2h.addWidget(self.txt_ara_hazir)
        t2h.addStretch()
        t2l.addLayout(t2h)
        self.tablo_hazir = QTableWidget(0,5)
        tablo_sag_tik_menu_ekle(self.tablo_hazir)
        self.tablo_hazir.setHorizontalHeaderLabels(["SIPARIS NO","MUSTERI","TERMIN","TOPLAM","ISLEM"])
        self.tablo_hazir.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tablo_hazir.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tablo_hazir.verticalHeader().setVisible(False)
        self.tablo_hazir.setShowGrid(False)
        self.tablo_hazir.setAlternatingRowColors(True)
        self.tablo_hazir.verticalHeader().setDefaultSectionSize(40)
        t2l.addWidget(self.tablo_hazir)
        self.tabs.addTab(t2, "Hazir Siparisler")

        # Sekme 3: Sevkiyat Geçmişi
        t3 = QWidget(); t3l = QVBoxLayout(t3); t3l.setContentsMargins(10,10,10,10); t3l.setSpacing(6)
        t3h = QHBoxLayout()
        self.txt_ara_sev = QLineEdit()
        self.txt_ara_sev.setPlaceholderText("🔍  Plaka, sofor veya siparis no ara...")
        self.txt_ara_sev.setFixedHeight(32)
        self.txt_ara_sev.setStyleSheet(
            "border:1.5px solid #dcdde1;border-radius:7px;padding:5px 10px;"
            "background:white;color:#2c3e50;font-size:12px;")
        self.txt_ara_sev.textChanged.connect(self._filtrele_sev)
        t3h.addWidget(self.txt_ara_sev)
        t3h.addStretch()
        t3l.addLayout(t3h)
        self.tablo_sev = QTableWidget(0,6)
        tablo_sag_tik_menu_ekle(self.tablo_sev)
        self.tablo_sev.setHorizontalHeaderLabels(["TARIH","PLAKA","SOFOR","TELEFON","SIPARISLER","DURUM"])
        self.tablo_sev.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tablo_sev.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tablo_sev.verticalHeader().setVisible(False)
        self.tablo_sev.setShowGrid(False)
        self.tablo_sev.setAlternatingRowColors(True)
        self.tablo_sev.verticalHeader().setDefaultSectionSize(40)
        self.tablo_sev.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tablo_sev.customContextMenuRequested.connect(self._sag_tik)
        t3l.addWidget(self.tablo_sev)
        self.tabs.addTab(t3, "Sevkiyat Gecmisi")

        lay.addWidget(self.tabs)

        # Sekme değişince yenile
        self.tabs.currentChanged.connect(lambda: self.yenile())

        # 30 saniyede bir otomatik yenile
        self._timer = QTimer(self)
        self._timer.timeout.connect(self.yenile)
        self._timer.start(30000)

    def hideEvent(self, event):
        self._timer.stop()
        super().hideEvent(event)

    def showEvent(self, event):
        self._timer.start(30000)
        super().showEvent(event)

    def _kart(self, b, v, r):
        f = QFrame(); f.setFixedSize(120,54)
        f.setStyleSheet(f"QFrame{{background:{r};border-radius:10px;border:none;}}")
        l = QVBoxLayout(f); l.setContentsMargins(8,4,8,4); l.setSpacing(0)
        lb = QLabel(b); lb.setStyleSheet("color:rgba(255,255,255,0.75);font-size:9px;font-weight:bold;background:transparent;letter-spacing:1px;")
        lv = QLabel(v); lv.setObjectName("Val"); lv.setStyleSheet("color:white;font-size:18px;font-weight:900;background:transparent;")
        l.addWidget(lb); l.addWidget(lv); return f

    def _set_kart(self, k, v):
        k.findChild(QLabel,"Val").setText(str(v))

    def _tablo_parca_doldur(self, filtre_sip_no=None):
        """
        Bekleyen parcalari siparis bazli gruplu goster.
        Ayni siparisten gelen parcalar renk gruplamasiyla birarada.
        filtre_sip_no: belirli bir siparisi filtrele (None=hepsi)
        """
        self.tablo_parca.setRowCount(0)
        self.btn_secili_sevk.setEnabled(False)

        try:
            # Unik siparisler
            self.cursor.execute("""
                SELECT DISTINCT sip_no, musteri
                FROM parca_sevk_bekliyor
                WHERE durum='Bekliyor'
                ORDER BY sip_no
            """)
            siparisler = self.cursor.fetchall()

            # Filtre dropdown guncelle
            self.cmb_sip_filtre.blockSignals(True)
            mevcut = self.cmb_sip_filtre.currentText()
            self.cmb_sip_filtre.clear()
            self.cmb_sip_filtre.addItem("Tum Siparisler")
            for row in siparisler:
                sno = row[0] if isinstance(row, (list,tuple)) else row.get("sip_no","")
                mus = row[1] if isinstance(row, (list,tuple)) else row.get("musteri","")
                self.cmb_sip_filtre.addItem("{} — {}".format(sno, mus))
            # Onceki secimi koru
            idx = self.cmb_sip_filtre.findText(mevcut)
            if idx >= 0:
                self.cmb_sip_filtre.setCurrentIndex(idx)
            self.cmb_sip_filtre.blockSignals(False)

            # Renk paleti — her siparis farkli arka plan
            GRUP_RENK = ["#EAF4FB","#EAFAF1","#FEF9E7","#F5EEF8","#FDEDEC","#EBF5FB"]

            satir = 0
            for g_idx, sip_row in enumerate(siparisler):
                sno = sip_row[0] if isinstance(sip_row,(list,tuple)) else sip_row.get("sip_no","")
                mus = sip_row[1] if isinstance(sip_row,(list,tuple)) else sip_row.get("musteri","")

                # Filtre uygula
                if filtre_sip_no and filtre_sip_no != "Tum Siparisler":
                    if "{} — {}".format(sno, mus) != filtre_sip_no:
                        continue

                grup_renk = GRUP_RENK[g_idx % len(GRUP_RENK)]

                # Grup baslik satiri
                self.tablo_parca.insertRow(satir)
                self.tablo_parca.setRowHeight(satir, 32)

                # Checkbox - grup basliginda yok
                self.tablo_parca.setCellWidget(satir, 0, None)

                baslik_it = QTableWidgetItem("  {} — {}".format(sno, mus))
                baslik_it.setFont(__import__('PyQt5.QtGui', fromlist=['QFont']).QFont(
                    "Segoe UI", 11, __import__('PyQt5.QtGui', fromlist=['QFont']).QFont.Bold))
                baslik_it.setBackground(QColor("#2C3E50"))
                baslik_it.setForeground(QColor("#FFFFFF"))
                baslik_it.setFlags(Qt.ItemIsEnabled)
                self.tablo_parca.setSpan(satir, 0, 1, 7)
                self.tablo_parca.setItem(satir, 0, baslik_it)
                satir += 1

                # Parcalari getir
                self.cursor.execute("""
                    SELECT id, parca_adi, bekleyen_adet, tamamlanan_adet, tarih
                    FROM parca_sevk_bekliyor
                    WHERE durum='Bekliyor' AND sip_no=?
                    ORDER BY id
                """, (sno,))
                parcalar = self.cursor.fetchall()

                for p_row in parcalar:
                    psb_id  = p_row[0] if isinstance(p_row,(list,tuple)) else p_row.get("id")
                    pad     = p_row[1] if isinstance(p_row,(list,tuple)) else p_row.get("parca_adi","")
                    bek_adet= p_row[2] if isinstance(p_row,(list,tuple)) else p_row.get("bekleyen_adet",0)
                    tam_adet= p_row[3] if isinstance(p_row,(list,tuple)) else p_row.get("tamamlanan_adet",0)
                    tarih   = p_row[4] if isinstance(p_row,(list,tuple)) else p_row.get("tarih","")

                    self.tablo_parca.insertRow(satir)
                    self.tablo_parca.setRowHeight(satir, 40)

                    # Checkbox
                    chk = QCheckBox()
                    chk.setStyleSheet("margin-left:10px;")
                    chk.setProperty("psb_id", psb_id)
                    chk.stateChanged.connect(self._secim_degisti)
                    self.tablo_parca.setCellWidget(satir, 0, chk)

                    # Veriler
                    vals = [sno, mus, pad or "-",
                            "{:g}".format(float(bek_adet or 0)),
                            "{:g}".format(float(tam_adet or 0)),
                            tarih or "-"]
                    for j, v in enumerate(vals):
                        it = QTableWidgetItem(v)
                        it.setTextAlignment(Qt.AlignCenter)
                        it.setData(Qt.UserRole, psb_id)
                        it.setBackground(QColor(grup_renk))
                        # Parca adi sola yasla
                        if j == 2:
                            it.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                        self.tablo_parca.setItem(satir, j+1, it)

                    satir += 1

        except Exception as e:
            print("tablo_parca_doldur HATA:", e)
            import traceback; traceback.print_exc()

    def _filtrele_parca(self, txt):
        """Parca tablosunda arama — grup baslik satirlarini gizleme."""
        txt = txt.lower().strip()
        for i in range(self.tablo_parca.rowCount()):
            # Grup baslik satiri mi? (span varsa)
            span = self.tablo_parca.columnSpan(i, 0)
            if span > 1:
                # Baslik satirini her zaman goster
                self.tablo_parca.setRowHidden(i, False)
                continue
            if not txt:
                self.tablo_parca.setRowHidden(i, False)
                continue
            # Tum sutunlarda ara
            gizle = True
            for c in range(1, self.tablo_parca.columnCount()):
                it = self.tablo_parca.item(i, c)
                if it and txt in it.text().lower():
                    gizle = False
                    break
            self.tablo_parca.setRowHidden(i, gizle)

    def _filtrele_hazir(self, txt):
        """Hazir siparisler tablosunda arama."""
        txt = txt.lower().strip()
        for i in range(self.tablo_hazir.rowCount()):
            if not txt:
                self.tablo_hazir.setRowHidden(i, False)
                continue
            gizle = True
            for c in range(self.tablo_hazir.columnCount()):
                it = self.tablo_hazir.item(i, c)
                if it and txt in it.text().lower():
                    gizle = False
                    break
            self.tablo_hazir.setRowHidden(i, gizle)

    def _filtrele_sev(self, txt):
        """Sevkiyat gecmisi tablosunda arama."""
        txt = txt.lower().strip()
        for i in range(self.tablo_sev.rowCount()):
            if not txt:
                self.tablo_sev.setRowHidden(i, False)
                continue
            gizle = True
            for c in range(self.tablo_sev.columnCount()):
                it = self.tablo_sev.item(i, c)
                if it and txt in it.text().lower():
                    gizle = False
                    break
            self.tablo_sev.setRowHidden(i, gizle)

    def _sip_filtre_degisti(self, text):
        """Filtre dropdown degisince tabloyu yenile."""
        if text == "Tum Siparisler":
            self._tablo_parca_doldur(None)
        else:
            self._tablo_parca_doldur(text)

    def _excel_export(self):
        if not excel_kaydet:
            return
        sutunlar = ["Plaka","Sofor","Tarih","Siparis Listesi","Durum"]
        satirlar = []
        try:
            self.cursor.execute("SELECT plaka, sofor, tarih, siparis_listesi, durum FROM sevkiyatlar ORDER BY id DESC")
            for row in self.cursor.fetchall():
                satirlar.append(list(row))
        except Exception as e:
            print("Sevkiyat excel hatasi:", e)
        excel_kaydet(self, "Sevkiyatlar", sutunlar, satirlar)

    def _wolvox_irsaliye(self):
        """
        Seçili sevkiyattaki parçaları Wolvox İrsaliye Excel şablonuna aktar.
        Seçili satır yoksa tüm bekleyen parçaları alır.
        """
        from PyQt5.QtWidgets import QFileDialog
        import os

        # Seçili parçaları topla
        parcalar = []

        # Önce tablo_parca'daki seçili/işaretli satırlara bak
        for i in range(self.tablo_parca.rowCount()):
            chk_widget = self.tablo_parca.cellWidget(i, 0)
            chk = chk_widget.findChild(QCheckBox) if chk_widget else None
            if chk and chk.isChecked():
                sip_no   = self.tablo_parca.item(i, 1).text() if self.tablo_parca.item(i, 1) else ""
                musteri  = self.tablo_parca.item(i, 2).text() if self.tablo_parca.item(i, 2) else ""
                parca    = self.tablo_parca.item(i, 3).text() if self.tablo_parca.item(i, 3) else ""
                adet     = self.tablo_parca.item(i, 4).text() if self.tablo_parca.item(i, 4) else "0"
                parcalar.append({
                    "sip_no":  sip_no,
                    "musteri": musteri,
                    "parca":   parca,
                    "adet":    adet,
                })

        # Hiç seçili yoksa tüm bekleyenleri al
        if not parcalar:
            try:
                self.cursor.execute("""
                    SELECT sip_no, musteri, parca_adi, bekleyen_adet
                    FROM parca_sevk_bekliyor WHERE durum='Bekliyor'
                    ORDER BY sip_no, id
                """)
                for row in self.cursor.fetchall():
                    parcalar.append({
                        "sip_no":  row[0] or "",
                        "musteri": row[1] or "",
                        "parca":   row[2] or "",
                        "adet":    str(row[3] or 0),
                    })
            except Exception as e:
                QMessageBox.warning(self, "Hata", "Parca verileri alinamadi:\n{}".format(e))
                return

        if not parcalar:
            QMessageBox.information(self, "Bilgi", "Sevk edilecek parça bulunamadı.")
            return

        # Kayıt yeri sor
        dosya, _ = QFileDialog.getSaveFileName(
            self, "Wolvox İrsaliye Dosyasını Kaydet",
            "Wolvox_Irsaliye_{}.xlsx".format(
                __import__('datetime').datetime.now().strftime("%Y%m%d_%H%M")),
            "Excel Dosyası (*.xlsx)")

        if not dosya:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Temp1"

            # Başlık satırı — Wolvox şablonu formatı
            basliklar = [
                "Sıra No", "Stok Kodu", "Stok Adı", "Miktarı", "Birimi",
                "Temel Mik.", "Temel Brm.", "Fiyatı", "Birim 2 Fiyatı",
                "KDV Durumu", "KDV Hrc.Fiyat", "Ara Tutarı", "KDV",
                "KDV siz Toplam", "KDV li Toplam", "Vade Günü", "Ek Bilgi 1", "Depo Adı"
            ]

            # Başlık stili
            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill("solid", start_color="2C3E50")
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin      = Side(style="thin", color="DDDDDD")
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, baslik in enumerate(basliklar, 1):
                cell = ws.cell(row=1, column=col, value=baslik)
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = hdr_align
                cell.border    = border

            ws.row_dimensions[1].height = 24

            # Sütun genişlikleri
            genislikler = [8, 14, 30, 10, 8, 10, 8, 12, 12, 10, 12, 12, 8, 14, 14, 10, 20, 14]
            for i, g in enumerate(genislikler, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = g

            # Satır stilleri
            alt_font   = Font(size=11)
            alt_align  = Alignment(horizontal="center", vertical="center")
            fill_tek   = PatternFill("solid", start_color="FFFFFF")
            fill_cift  = PatternFill("solid", start_color="F8F9FA")

            # Veri satırları
            for idx, p in enumerate(parcalar, 1):
                row = idx + 1
                fill = fill_tek if idx % 2 == 1 else fill_cift

                degerler = [
                    idx,           # Sıra No
                    "",            # Stok Kodu — Wolvox'ta elle eşleştirilir
                    p["parca"],    # Stok Adı
                    float(p["adet"]) if p["adet"] else 1,  # Miktarı
                    "Adet",        # Birimi
                    "",            # Temel Mik.
                    "",            # Temel Brm.
                    0,             # Fiyatı
                    0,             # Birim 2 Fiyatı
                    "Dahil",       # KDV Durumu
                    0,             # KDV Hrc.Fiyat
                    0,             # Ara Tutarı
                    20,            # KDV (%)
                    0,             # KDV siz Toplam
                    0,             # KDV li Toplam
                    0,             # Vade Günü
                    "{} - {}".format(p["sip_no"], p["musteri"]),  # Ek Bilgi 1
                    "Merkez",      # Depo Adı
                ]

                for col, val in enumerate(degerler, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font      = alt_font
                    cell.alignment = alt_align
                    cell.fill      = fill
                    cell.border    = border

                ws.row_dimensions[row].height = 22

            wb.save(dosya)

            cevap = QMessageBox.question(
                self, "✅ Başarılı",
                "{} parca Wolvox formatina aktarildi!\n\nDosya: {}\n\nDosyayi simdi acmak ister misiniz?".format(
                    len(parcalar), dosya),
                QMessageBox.Yes | QMessageBox.No)

            if cevap == QMessageBox.Yes:
                try:
                    import subprocess, sys
                    if sys.platform == "win32":
                        os.startfile(dosya)
                    else:
                        subprocess.Popen(["xdg-open", dosya])
                except:
                    pass

        except ImportError:
            QMessageBox.critical(self, "Hata",
                "openpyxl kutuphanesi bulunamadi!\n\nKomut istemcisinde sunu calistirin:\npip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Hata", "Irsaliye olusturulamadi:\n{}".format(e))

    def yenile(self):
        try:
            # Kartlar
            self.cursor.execute(
                "SELECT COUNT(*) FROM parca_sevk_bekliyor WHERE durum='Bekliyor'")
            self._set_kart(self.k_parca, self.cursor.fetchone()[0])
            self.cursor.execute(
                "SELECT COUNT(*) FROM siparisler WHERE durum='Hazir'")
            self._set_kart(self.k_hazir, self.cursor.fetchone()[0])
            self.cursor.execute(
                "SELECT COUNT(*) FROM sevkiyatlar WHERE durum='Yolda'")
            self._set_kart(self.k_yolda, self.cursor.fetchone()[0])
            self.cursor.execute(
                "SELECT COUNT(*) FROM sevkiyatlar WHERE durum='Teslim Edildi'")
            self._set_kart(self.k_teslim, self.cursor.fetchone()[0])

            # Sekme 1: Uretimden Gelenler — siparis bazli gruplu
            self._tablo_parca_doldur()

            # Sekme 2: Hazır Siparişler
            self.cursor.execute("""
                SELECT id, sip_no, musteri, termin, genel_toplam
                FROM siparisler WHERE durum='Hazir' ORDER BY id DESC
            """)
            self.tablo_hazir.setRowCount(0)
            for i, (sid, sno, mus, ter, top) in enumerate(self.cursor.fetchall()):
                self.tablo_hazir.insertRow(i)
                for j, v in enumerate([sno, mus or "-", ter or "-",
                                        "{:,.2f}".format(float(top or 0))]):
                    it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter)
                    it.setData(Qt.UserRole, sid)
                    self.tablo_hazir.setItem(i, j, it)
                btn = QPushButton("Sevke Al"); btn.setFixedHeight(32); btn.setMinimumWidth(90)
                btn.setStyleSheet(
                    "background:#e67e22;color:white;font-weight:bold;"
                    "font-size:12px;border-radius:6px;border:none;padding:4px 12px;")
                btn.clicked.connect(lambda _, s=sid: self._tek_sevk(s))
                bw = QWidget(); bl = QHBoxLayout(bw)
                bl.setContentsMargins(4,4,4,4); bl.addWidget(btn)
                self.tablo_hazir.setCellWidget(i, 4, bw)

            # Sekme 3: Sevkiyat Geçmişi
            self.cursor.execute("""
                SELECT id, tarih, plaka, sofor, telefon, siparis_listesi, durum
                FROM sevkiyatlar ORDER BY id DESC
            """)
            self.tablo_sev.setRowCount(0)
            for i, (sev_id, tarih, plaka, sofor, tel, sip_l, durum) in enumerate(
                    self.cursor.fetchall()):
                self.cursor.execute(
                    "SELECT COUNT(*) FROM sevkiyat_siparisler WHERE sevkiyat_id=?",
                    (sev_id,))
                sip_sayi = self.cursor.fetchone()[0]
                self.tablo_sev.insertRow(i)
                renk = {"Yolda":"#e67e22","Teslim Edildi":"#27ae60"}.get(durum,"#7f8c8d")
                for j, v in enumerate([tarih or "-", plaka or "-", sofor or "-",
                                        tel or "-", "{} siparis".format(sip_sayi), durum or "-"]):
                    it = QTableWidgetItem(v); it.setTextAlignment(Qt.AlignCenter)
                    it.setData(Qt.UserRole, sev_id)
                    if j == 5: it.setForeground(QColor(renk))
                    self.tablo_sev.setItem(i, j, it)

        except Exception as e:
            print("Sevkiyat yenile hatasi:", e)

    def _secim_degisti(self):
        """Herhangi bir checkbox değişince Sevk Et butonunu aktif/pasif yap."""
        secili_var = any(
            self.tablo_parca.cellWidget(r, 0) and
            self.tablo_parca.cellWidget(r, 0).isChecked()
            for r in range(self.tablo_parca.rowCount())
        )
        self.btn_secili_sevk.setEnabled(secili_var)

    def _tumunu_sec(self):
        """Tüm satırları seç / seçimi kaldır."""
        tumu_secili = all(
            self.tablo_parca.cellWidget(r, 0) and
            self.tablo_parca.cellWidget(r, 0).isChecked()
            for r in range(self.tablo_parca.rowCount())
            if self.tablo_parca.cellWidget(r, 0)
        )
        for r in range(self.tablo_parca.rowCount()):
            chk = self.tablo_parca.cellWidget(r, 0)
            if chk: chk.setChecked(not tumu_secili)

    def _secilileri_sevk_et(self):
        """Seçili parçaları tek dialog ile toplu sevk et."""
        secili_ids = []
        secili_adlar = []
        for r in range(self.tablo_parca.rowCount()):
            chk = self.tablo_parca.cellWidget(r, 0)
            if chk and chk.isChecked():
                it = self.tablo_parca.item(r, 1)  # sip_no sütunu
                if it:
                    secili_ids.append(it.data(Qt.UserRole))
                    pad = self.tablo_parca.item(r, 3)
                    secili_adlar.append(pad.text() if pad else "-")

        if not secili_ids:
            QMessageBox.warning(self, "Uyari", "En az bir parca secin."); return

        # Müşteri bilgisini al (ilk seçiliden)
        musteri_it = self.tablo_parca.item(
            [r for r in range(self.tablo_parca.rowCount())
             if self.tablo_parca.cellWidget(r,0) and
             self.tablo_parca.cellWidget(r,0).isChecked()][0], 2)
        musteri = musteri_it.text() if musteri_it else "-"

        ozet = "{} parca".format(len(secili_ids))
        dlg = ParcaSevkDialog(ozet, musteri, self)
        if dlg.exec_() != QDialog.Accepted:
            return

        try:
            tarih = datetime.now().strftime("%d.%m.%Y %H:%M")
            parca_listesi = ", ".join(secili_adlar)
            self.cursor.execute("""
                INSERT INTO sevkiyatlar
                    (plaka, sofor, telefon, tarih, siparis_listesi, notlar, durum)
                VALUES (?,?,?,?,?,?,'Yolda')
            """, (dlg.plaka, dlg.sofor, dlg.telefon, tarih,
                  parca_listesi, ""))

            for psb_id in secili_ids:
                self.cursor.execute(
                    "UPDATE parca_sevk_bekliyor SET durum='Sevk Edildi' WHERE id=?",
                    (psb_id,))

            self.conn.commit()

            # Sipariş senkronizasyonu — bekleyen parça kalmadıysa siparişi kapat
            self._siparis_durum_guncelle(secili_ids)

            log_yaz(self.cursor, self.conn, "TOPLU_PARCA_SEVK",
                    "{} parca | {} | {}".format(len(secili_ids), dlg.plaka, dlg.sofor))
            QMessageBox.information(
                self, "Sevk Edildi",
                "{} parca sevk edildi.\n\nArac: {}\nSofor: {}".format(
                    len(secili_ids), dlg.plaka, dlg.sofor))
            self.yenile()
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))

    def _siparis_durum_guncelle(self, sevk_edilen_psb_ids):
        """Sevk edilen parçaların sipariş ID'lerini bul,
        siparişi 'Sevk Edildi' yap."""
        try:
            if not sevk_edilen_psb_ids: return
            ph = ",".join("?" * len(sevk_edilen_psb_ids))

            # UPDATE öncesi siparis_id'leri al (id listesi ile)
            self.cursor.execute(
                "SELECT DISTINCT siparis_id FROM parca_sevk_bekliyor "
                "WHERE id IN ({}) AND siparis_id IS NOT NULL".format(ph),
                sevk_edilen_psb_ids)
            sip_idler = [r[0] for r in self.cursor.fetchall()]

            print("Senkronizasyon - siparis idleri:", sip_idler)

            for sip_id in sip_idler:
                self.cursor.execute(
                    "UPDATE siparisler SET durum='Sevk Edildi' "
                    "WHERE id=? AND durum NOT IN ('Iptal', 'Faturalandı')",
                    (sip_id,))
                print("Siparis {} -> Sevk Edildi guncellendi".format(sip_id))

            self.conn.commit()
        except Exception as e:
            print("Siparis durum guncelleme hatasi:", e)
            import traceback; traceback.print_exc()

    def _yeni_sevk(self):
        dlg = YeniSevkDialog(self.cursor, self.conn, self)
        if dlg.exec_() == QDialog.Accepted:
            self.yenile()

    def _tek_sevk(self, sid):
        plaka, ok = QInputDialog.getText(self,"Sevke Al","Araç plakası:")
        if not ok or not plaka: return
        sofor, ok2 = QInputDialog.getText(self,"Sevke Al","Şoför adı:")
        if not ok2: return
        try:
            tarih = datetime.now().strftime("%d.%m.%Y %H:%M")
            self.cursor.execute("""
                INSERT INTO sevkiyatlar (plaka, sofor, tarih, siparis_listesi, durum)
                VALUES (?,?,?,?,'Yolda')
            """, (plaka, sofor, tarih, str(sid)))
            sev_id = self.cursor.lastrowid
            self.cursor.execute("INSERT INTO sevkiyat_siparisler (sevkiyat_id, siparis_id) VALUES (?,?)", (sev_id, sid))
            self.cursor.execute("UPDATE siparisler SET durum='Sevk Edildi', arac=?, sofor=? WHERE id=?", (plaka, sofor, sid))
            self.conn.commit()
            log_yaz(self.cursor, self.conn, "SEVKIYAT_OLUSTURULDU", f"{plaka} | {sofor} | Sipariş ID:{sid}")
            self.yenile()
        except Exception as e:
            QMessageBox.critical(self,"Hata",str(e))

    def _sag_tik(self, pos):
        row = self.tablo_sev.currentRow()
        if row < 0: return
        item = self.tablo_sev.item(row, 0)
        sev_id = item.data(Qt.UserRole) if item else None
        if not sev_id: return
        menu = QMenu(self)
        menu.setStyleSheet("QMenu{background:white;border:1px solid #dcdde1;border-radius:8px;padding:6px;font-size:13px;} QMenu::item{padding:8px 18px;border-radius:4px;} QMenu::item:selected{background:#fde8e8;color:#c0392b;}")
        act_teslim = menu.addAction("✅ Teslim Edildi Olarak İşaretle")
        secim = menu.exec_(self.tablo_sev.mapToGlobal(pos))
        if secim == act_teslim:
            self.cursor.execute("UPDATE sevkiyatlar SET durum='Teslim Edildi' WHERE id=?", (sev_id,))
            self.conn.commit()
            log_yaz(self.cursor, self.conn, "SEVKIYAT_TESLIM", f"Sevkiyat ID:{sev_id}")
            self.yenile()
